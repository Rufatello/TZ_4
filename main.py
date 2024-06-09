from tkinter import *
from tkinter import ttk
import openpyxl

file = openpyxl.load_workbook('Задание 4.xlsx')
sheet = file.active

dataset = []

for i in range(2, sheet.max_row + 1):
    numb = sheet.cell(row=i, column=1).value
    par_1 = sheet.cell(row=i, column=2).value
    par_2 = sheet.cell(row=i, column=3).value
    par_3 = sheet.cell(row=i, column=4).value

    if par_3 is not None:
        par_3 = round(par_3, 2)
    par_4 = sheet.cell(row=i, column=5).value
    if par_4 is not None:
        par_4 = round(par_4, 2)
    par_5 = sheet.cell(row=i, column=6).value
    if par_5 is not None:
        par_5 = round(par_5, 2)
    par_6 = sheet.cell(row=i, column=7).value
    if par_6 is not None:
        par_6 = round(par_6, 1)
    par_7 = sheet.cell(row=i, column=8).value
    if par_7 is not None:
        par_7 = round(par_7, 2)

    if numb is not None and par_1 is not None and par_2 is not None and par_3 is not None and par_4 is not None and par_5 is not None and par_6 is not None and par_7 is not None:
        dataset.append((numb, par_1, par_2, par_3, par_4, par_5, par_6, par_7))
    else:
        break

root = Tk()

root.title('Северсталь')
root.geometry('900x900')

label = Label(text='Выберите ')
columns = ("№ партии", "Параметр_1", "Параметр_2", "Параметр_3", "Параметр_4", "Параметр_5", "Параметр_6", "Параметр_7")
tree = ttk.Treeview(columns=columns, show="headings")
tree.pack(side=TOP, fill=BOTH, expand=1)

tree.heading("№ партии", text="№ партии")
tree.heading("Параметр_1", text="Параметр_1")
tree.heading("Параметр_2", text="Параметр_2")
tree.heading("Параметр_3", text="Параметр_3")
tree.heading("Параметр_4", text="Параметр_4")
tree.heading("Параметр_5", text="Параметр_5")
tree.heading("Параметр_6", text="Параметр_6")
tree.heading("Параметр_7", text="Параметр_7")

tree.column("#1", stretch=NO, width=90)
tree.column("#2", stretch=NO, width=110)
tree.column("#3", stretch=NO, width=110)
tree.column("#4", stretch=NO, width=110)
tree.column("#5", stretch=NO, width=110)
tree.column("#6", stretch=NO, width=110)
tree.column("#7", stretch=NO, width=100)

for data in dataset:
    tree.insert("", END, values=data)
label.pack()
data = []


def item_selected(event):
    global data
    data.clear()
    for selected_item in tree.selection():
        item = tree.item(selected_item)
        param_1 = item["values"][1]
        param_2 = item["values"][2]
        param_3 = float(item["values"][3])
        param_4 = float(item["values"][4])
        param_5 = float(item["values"][5])
        param_6 = float(item["values"][6])
        param_7 = float(item["values"][7])

        for i in dataset:
            if param_1 == i[1] and abs(i[2] - param_2) <= 10 and abs(i[3] - param_3) <= 0.8 and abs(
                    i[4] - param_4) <= 1.5 and abs(i[5] - param_5) <= 1 \
                    and abs(i[6] - param_6) <= 9 \
                    and abs(i[7] - param_7) <= 1:
                data.append(i)

    selected_data = ""

    for item in set(data):
        selected_data += f"{item}\n"

    label["text"] = selected_data


tree.bind("<<TreeviewSelect>>", item_selected)

root.mainloop()
